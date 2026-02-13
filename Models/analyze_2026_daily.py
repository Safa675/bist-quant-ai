#!/usr/bin/env python3
"""
Build daily factor-vs-benchmark return comparison for a single year.

Uses existing backtest artifacts in `Models/results/*/returns.csv`.
No signal/backtest recomputation is performed.

Default behavior compares all factor folders against `xu100` for year 2026.
Also attempts to add XU030 benchmark from Yahoo Finance (`XU030.IS`).
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, Tuple

import numpy as np
import pandas as pd

try:
    import yfinance as yf
except Exception:
    yf = None

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


WIN_MARK = "✅"
LOSE_MARK = "❌"
FLAT_MARK = "➖"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare all factor daily returns vs benchmarks using results CSV files.",
    )
    parser.add_argument(
        "--results-dir",
        type=str,
        default=str(Path(__file__).parent / "results"),
        help="Path to results directory (default: Models/results)",
    )
    parser.add_argument(
        "--year",
        type=int,
        default=2026,
        help="Calendar year to analyze (default: 2026)",
    )
    parser.add_argument(
        "--benchmark",
        type=str,
        default="xu100",
        help="Benchmark folder name under results dir (default: xu100)",
    )
    parser.add_argument(
        "--xu030-symbol",
        type=str,
        default="XU030.IS",
        help="Yahoo Finance symbol for BIST 30 benchmark (default: XU030.IS)",
    )
    parser.add_argument(
        "--skip-xu030",
        action="store_true",
        help="Skip Yahoo Finance XU030 download even if available.",
    )
    return parser.parse_args()


def _load_returns_series(returns_path: Path, series_name: str) -> pd.Series:
    """Load standardized daily return series from one returns.csv file."""
    df = pd.read_csv(returns_path)
    if "date" not in df.columns:
        raise ValueError(f"'date' column not found in {returns_path}")
    if "Return" not in df.columns:
        raise ValueError(f"'Return' column not found in {returns_path}")

    df = df[["date", "Return"]].copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["Return"] = pd.to_numeric(df["Return"], errors="coerce")
    df = df.dropna(subset=["date"])
    df = df.sort_values("date")
    df = df.drop_duplicates(subset=["date"], keep="last")

    series = df.set_index("date")["Return"].astype(float)
    series.name = series_name
    return series


def _extract_close_from_yf_download(df: pd.DataFrame) -> pd.Series:
    """Extract close-like series from yfinance download output."""
    if df.empty:
        return pd.Series(dtype=float)

    if isinstance(df.columns, pd.MultiIndex):
        for field in ("Adj Close", "Close"):
            if field in df.columns.get_level_values(0):
                sub = df[field]
                if isinstance(sub, pd.DataFrame):
                    return pd.to_numeric(sub.iloc[:, 0], errors="coerce")
                return pd.to_numeric(sub, errors="coerce")
    else:
        for field in ("Adj Close", "Close"):
            if field in df.columns:
                return pd.to_numeric(df[field], errors="coerce")

    return pd.Series(dtype=float)


def _load_xu030_from_yfinance(symbol: str, year: int) -> Tuple[pd.Series | None, str | None]:
    """Load XU030 returns from yfinance for the given year."""
    if yf is None:
        return None, "yfinance is not installed"

    start = pd.Timestamp(year=year, month=1, day=1)
    end = pd.Timestamp(year=year + 1, month=1, day=1)

    try:
        raw = yf.download(
            symbol,
            start=start,
            end=end,
            auto_adjust=False,
            progress=False,
            threads=False,
        )
    except Exception as exc:
        return None, f"yfinance download failed for {symbol}: {exc}"

    close = _extract_close_from_yf_download(raw).dropna()
    if close.empty:
        return None, f"no price data returned for {symbol}"

    close.index = pd.to_datetime(close.index, errors="coerce")
    close = close[close.index.notna()].sort_index()
    returns = close.pct_change(fill_method=None).dropna()
    returns = returns[returns.index.year == year]
    if returns.empty:
        return None, f"no {year} returns for {symbol}"

    returns.name = "XU030_Return"
    return returns, None


def _annualized_tracking_error(excess: pd.Series) -> float:
    if len(excess) < 2:
        return np.nan
    return float(excess.std(ddof=1) * np.sqrt(252))


def _information_ratio(excess: pd.Series) -> float:
    if len(excess) < 2:
        return np.nan
    std = excess.std(ddof=1)
    if std <= 0 or pd.isna(std):
        return np.nan
    return float((excess.mean() / std) * np.sqrt(252))


def _status_marks(excess: pd.Series) -> pd.Series:
    """Map excess returns to check/cross/flat marks."""
    status = pd.Series("", index=excess.index, dtype=object)
    status = status.where(~(excess > 0), WIN_MARK)
    status = status.where(~(excess < 0), LOSE_MARK)
    status = status.where(~(excess == 0), FLAT_MARK)
    status = status.where(excess.notna(), "")
    return status


def _safe_sheet_name(raw_name: str, used: set[str]) -> str:
    """Excel-safe unique sheet name."""
    cleaned = raw_name
    for ch in ("\\", "/", "*", "[", "]", ":", "?"):
        cleaned = cleaned.replace(ch, "_")
    cleaned = cleaned[:31] if cleaned else "Sheet"

    candidate = cleaned
    suffix = 1
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{cleaned[:31-len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def _write_excel_report(
    out_path: Path,
    per_factor_daily: Dict[str, pd.DataFrame],
    summary: pd.DataFrame,
) -> None:
    """Write one workbook with summary + one sheet per factor."""
    used_sheet_names: set[str] = set()
    factor_sheet_names: dict[str, str] = {}

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False)
        used_sheet_names.add("summary")

        for factor, df in sorted(per_factor_daily.items()):
            sheet_name = _safe_sheet_name(factor, used_sheet_names)
            factor_sheet_names[factor] = sheet_name
            out_df = df.reset_index().rename(columns={"index": "date"})
            out_df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(out_path)

    # Summary sheet formatting
    ws_summary = wb["summary"]
    ws_summary.freeze_panes = "A2"
    ws_summary.auto_filter.ref = ws_summary.dimensions

    summary_percent_keywords = ("Rate", "Return", "Excess", "Tracking_Error")
    for col_idx, cell in enumerate(ws_summary[1], start=1):
        header = str(cell.value) if cell.value is not None else ""
        if any(k in header for k in summary_percent_keywords):
            for row in range(2, ws_summary.max_row + 1):
                ws_summary.cell(row=row, column=col_idx).number_format = "0.00%"

    # Factor sheet formatting (including win/loss marks)
    green_font = Font(color="FF008000", bold=True)
    red_font = Font(color="FFFF0000", bold=True)
    gray_font = Font(color="FF808080")

    for factor, sheet_name in factor_sheet_names.items():
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
        for col_idx, header in enumerate(headers, start=1):
            if (
                "Return" in header
                or "Excess" in header
            ) and header not in ("vs_XU100", "vs_XU030"):
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = "0.00%"

            if header in ("vs_XU100", "vs_XU030"):
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=col_idx)
                    if c.value == WIN_MARK:
                        c.font = green_font
                    elif c.value == LOSE_MARK:
                        c.font = red_font
                    elif c.value == FLAT_MARK:
                        c.font = gray_font

        # Basic width tuning for readability.
        for col_idx, header in enumerate(headers, start=1):
            if header == "date":
                width = 12
            elif header.startswith("vs_"):
                width = 11
            else:
                width = max(14, min(28, len(header) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(out_path)


def main() -> None:
    args = parse_args()
    results_dir = Path(args.results_dir)
    year = int(args.year)
    benchmark_name = str(args.benchmark)

    benchmark_path = results_dir / benchmark_name / "returns.csv"
    if not benchmark_path.exists():
        raise FileNotFoundError(f"Benchmark file not found: {benchmark_path}")

    benchmark = _load_returns_series(benchmark_path, "XU100_Return")
    benchmark = benchmark[benchmark.index.year == year]
    if benchmark.empty:
        raise ValueError(f"No benchmark rows found for year {year} in {benchmark_path}")

    xu030 = None
    if not args.skip_xu030:
        xu030, xu030_err = _load_xu030_from_yfinance(args.xu030_symbol, year)
        if xu030 is None:
            print(f"XU030 unavailable ({args.xu030_symbol}): {xu030_err}")
        else:
            print(f"Loaded XU030 benchmark from Yahoo ({args.xu030_symbol}) with {len(xu030)} rows")

    factor_returns: Dict[str, pd.Series] = {}
    for factor_dir in sorted([p for p in results_dir.iterdir() if p.is_dir()]):
        factor = factor_dir.name
        if factor == benchmark_name:
            continue
        returns_path = factor_dir / "returns.csv"
        if not returns_path.exists():
            continue

        try:
            series = _load_returns_series(returns_path, factor)
        except Exception as exc:
            print(f"Skipping {factor}: {exc}")
            continue

        series = series[series.index.year == year]
        factor_returns[factor] = series

    if not factor_returns:
        raise ValueError(f"No factor return series found for year {year} under {results_dir}")

    # Wide daily matrix: date + benchmarks + factor returns/excess.
    daily_wide = pd.DataFrame(index=benchmark.index)
    daily_wide["XU100_Return"] = benchmark
    if xu030 is not None:
        daily_wide["XU030_Return"] = xu030.reindex(daily_wide.index)

    long_frames = []
    summary_rows = []
    per_factor_daily: Dict[str, pd.DataFrame] = {}

    for factor, series in factor_returns.items():
        aligned_xu100 = pd.concat(
            [series.rename("Factor_Return"), benchmark.rename("XU100_Return")],
            axis=1,
            join="inner",
        ).dropna()
        if aligned_xu100.empty:
            continue

        factor_daily = aligned_xu100.copy()
        factor_daily["Excess_vs_XU100"] = factor_daily["Factor_Return"] - factor_daily["XU100_Return"]
        factor_daily["vs_XU100"] = _status_marks(factor_daily["Excess_vs_XU100"])

        if xu030 is not None:
            factor_daily["XU030_Return"] = xu030.reindex(factor_daily.index)
            factor_daily["Excess_vs_XU030"] = factor_daily["Factor_Return"] - factor_daily["XU030_Return"]
            factor_daily["vs_XU030"] = _status_marks(factor_daily["Excess_vs_XU030"])
        else:
            factor_daily["XU030_Return"] = np.nan
            factor_daily["Excess_vs_XU030"] = np.nan
            factor_daily["vs_XU030"] = ""

        per_factor_daily[factor] = factor_daily

        daily_wide[f"{factor}_Return"] = series.reindex(daily_wide.index)
        daily_wide[f"{factor}_Excess_vs_XU100"] = daily_wide[f"{factor}_Return"] - daily_wide["XU100_Return"]
        if xu030 is not None:
            daily_wide[f"{factor}_Excess_vs_XU030"] = daily_wide[f"{factor}_Return"] - daily_wide["XU030_Return"]

        long_df = factor_daily.copy()
        long_df.insert(0, "Factor", factor)
        long_frames.append(long_df.reset_index().rename(columns={"index": "date"}))

        excess_xu100 = factor_daily["Excess_vs_XU100"].dropna()
        factor_cum = float((1.0 + aligned_xu100["Factor_Return"]).prod() - 1.0)
        bench_xu100_cum = float((1.0 + aligned_xu100["XU100_Return"]).prod() - 1.0)
        relative_cum_xu100 = float((1.0 + factor_cum) / (1.0 + bench_xu100_cum) - 1.0)

        row = {
            "Factor": factor,
            "Obs_Days": int(len(aligned_xu100)),
            "Avg_Daily_Return": float(aligned_xu100["Factor_Return"].mean()),
            "Avg_Daily_XU100_Return": float(aligned_xu100["XU100_Return"].mean()),
            "Avg_Daily_Excess_vs_XU100": float(excess_xu100.mean()),
            "Outperform_Days_vs_XU100": int((excess_xu100 > 0).sum()),
            "Outperform_Rate_vs_XU100": float((excess_xu100 > 0).mean()),
            "Cumulative_Return": factor_cum,
            "XU100_Cumulative_Return": bench_xu100_cum,
            "Relative_Cumulative_Return_vs_XU100": relative_cum_xu100,
            "Tracking_Error_Annualized_vs_XU100": _annualized_tracking_error(excess_xu100),
            "Information_Ratio_vs_XU100": _information_ratio(excess_xu100),
        }

        if xu030 is not None:
            aligned_xu030 = factor_daily[["Factor_Return", "XU030_Return"]].dropna()
            if not aligned_xu030.empty:
                excess_xu030 = aligned_xu030["Factor_Return"] - aligned_xu030["XU030_Return"]
                xu030_cum = float((1.0 + aligned_xu030["XU030_Return"]).prod() - 1.0)
                relative_cum_xu030 = float((1.0 + factor_cum) / (1.0 + xu030_cum) - 1.0)
                row["Obs_Days_vs_XU030"] = int(len(aligned_xu030))
                row["Avg_Daily_XU030_Return"] = float(aligned_xu030["XU030_Return"].mean())
                row["Avg_Daily_Excess_vs_XU030"] = float(excess_xu030.mean())
                row["Outperform_Days_vs_XU030"] = int((excess_xu030 > 0).sum())
                row["Outperform_Rate_vs_XU030"] = float((excess_xu030 > 0).mean())
                row["XU030_Cumulative_Return"] = xu030_cum
                row["Relative_Cumulative_Return_vs_XU030"] = relative_cum_xu030
                row["Tracking_Error_Annualized_vs_XU030"] = _annualized_tracking_error(excess_xu030)
                row["Information_Ratio_vs_XU030"] = _information_ratio(excess_xu030)
            else:
                row["Obs_Days_vs_XU030"] = 0
                row["Avg_Daily_XU030_Return"] = np.nan
                row["Avg_Daily_Excess_vs_XU030"] = np.nan
                row["Outperform_Days_vs_XU030"] = 0
                row["Outperform_Rate_vs_XU030"] = np.nan
                row["XU030_Cumulative_Return"] = np.nan
                row["Relative_Cumulative_Return_vs_XU030"] = np.nan
                row["Tracking_Error_Annualized_vs_XU030"] = np.nan
                row["Information_Ratio_vs_XU030"] = np.nan

        summary_rows.append(row)

    if not summary_rows:
        raise ValueError(f"No overlapping daily rows between factors and benchmark for year {year}")

    daily_wide = daily_wide.sort_index()
    daily_long = pd.concat(long_frames, ignore_index=True).sort_values(["date", "Factor"])
    summary = pd.DataFrame(summary_rows).sort_values(
        ["Relative_Cumulative_Return_vs_XU100", "Avg_Daily_Excess_vs_XU100"],
        ascending=False,
    )

    out_workbook = results_dir / f"daily_vs_{benchmark_name}_{year}_by_signal.xlsx"

    _write_excel_report(out_workbook, per_factor_daily, summary)

    print("=" * 72)
    print(f"DAILY FACTOR RETURN COMPARISON VS {benchmark_name.upper()} (+XU030) ({year})")
    print("=" * 72)
    print(f"Benchmark days in scope: {len(benchmark)}")
    print(f"Factors compared: {len(summary)}")
    print(f"Date range: {daily_wide.index.min().date()} -> {daily_wide.index.max().date()}")
    print("")
    print("Top 10 by relative cumulative return:")
    top_cols = [
        "Factor",
        "Obs_Days",
        "Relative_Cumulative_Return_vs_XU100",
        "Avg_Daily_Excess_vs_XU100",
        "Outperform_Rate_vs_XU100",
        "Information_Ratio_vs_XU100",
    ]
    print(summary[top_cols].head(10).to_string(index=False))
    print("")
    print(f"Saved: {out_workbook}")


if __name__ == "__main__":
    main()
